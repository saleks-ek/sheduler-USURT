'''
Формирует расписание из файлов расположенных а папке "Расписания" и создает файл Расписание.xslx
вне папки "Расписания".
'''
import os
import xlrd
import openpyxl as opx
from openpyxl.styles import Font

os.chdir('Расписания\\')
fileList = os.listdir()
pass
sheduleList: list[str] = list()
kabinetSet:set[str] = set()
groupSet:set[str] = set()
edukatorSet:set[str] = set()
fileNameXLSX: str = 'Расписание.xlsx' 
timeList = ['08:30-10:05','10:15-11:50','12:00-13:35','14:30-16:05','16:15-17:50','18:00-19:35','19:45-21:20']
month = {'01':'января', '02':'февраля', '03':'марта', '04':'апреля', '05':'мая', '06':'июня', '07':'июля', '10':'октября','11':'ноября','12':'декабря'}
dataList: list[str] = list()
dayWeekList: list[str] = list()
dataDayWeekSet:set[str] = set()
kabinetBusyForCount: list[str] = list()

def readFileWriteList(FileXLS,sheduleList,kabinetSet,edukatorSet,kabinetBusyForCount):
    fileNameXLS: str = FileXLS # 'ФУПП 1 курс  нечетная.xls' # ФУПП 1 курс  нечетная.xls

    wb = xlrd.open_workbook(fileNameXLS)
    sheet = wb.sheet_by_index(0)
    sheduleListString = list()
    rowGroup: int = 2 # строка со именами групп

    # print("Максимальный номер строки: ", sheet.nrows)
    # print("Максимальный номер столбца: ",sheet.ncols)

    faculty: str = sheet.cell_value(0,0)
    period: str = sheet.cell_value(1,0)
    semestr: str = period[0:9]
    # year: str = period[10:19]
    week: str = period[20:]
    pass
    for row in range(3, sheet.nrows):
        for column in range(0, sheet.ncols):
            sheduleListString = []
            cellString = sheet.cell_value(row, column)
            if column == 0:
                if cellString != '':
                    day = cellString
                    if week == 'Нечетная' or week == 'Четная':
                        data = day[:day.find('\n')]
                        dayWeek = day[day.find('\n')+1:]
                        dataDayWeekSet.add(day)
                    else:
                        dayWeek = day[:day.find('\n')]
                        data = day[day.find('\n')+1:]
                        number = data[:data.find('.')]
                        data = data[data.find('.')+1:]
                        monthDay = data[:data.find('.')]
                        # monthDay = month[monthDay]
                        data = number + ' ' + month[monthDay]
                        day = data + '\n' + dayWeek
                        dataDayWeekSet.add(day)
            elif column == 1:
                if cellString != '':
                    time = cellString
            else:
                group = sheet.cell_value(rowGroup, column)
                point = group.find('-')+1
                course = group[point:point+1]
                if cellString != '\n':
                    lessonString = cellString
                    if  lessonString.find('\n') != -1:
                        discipline: str = lessonString[1:lessonString.find('\n')]
                        if discipline[0:2] == '  ':
                            discipline = discipline[2:]
                        if lessonString[lessonString.find('\n')+1:lessonString.find('\n')+4] == '-  ':
                            lessonString = lessonString[lessonString.find('\n')+4:]
                        else: lessonString = lessonString[lessonString.find('\n')+3:]
                        edukator = lessonString[:lessonString.find(',')]
                        lessonString = lessonString[lessonString.find('\n')+4:]
                        kabinet = lessonString[:lessonString.find(',')]
                        typeLesson = lessonString[lessonString.find(',')+2:]
                        sheduleListString.append(faculty) # 0
                        sheduleListString.append(course) # 1
                        sheduleListString.append(semestr) # 2
                        sheduleListString.append(week) # 3
                        sheduleListString.append(data) # 4
                        sheduleListString.append(dayWeek) # 5
                        sheduleListString.append(time) # 6
                        sheduleListString.append(group) # 7
                        sheduleListString.append(discipline) # 8
                        sheduleListString.append(edukator) # 9 
                        sheduleListString.append(kabinet) # 10
                        kabinetBusyForCount.append(kabinet)
                        sheduleListString.append(typeLesson) # 11
                        kabinetSet.add(kabinet)
                        # groupSet.add(group)
                        edukatorSet.add(edukator)
                        sheduleList.append(sheduleListString)
    print('*', end='')
    return(sheduleList,kabinetSet,edukatorSet,dataDayWeekSet)

for file in fileList:
    readFileWriteList(file,sheduleList,kabinetSet,edukatorSet,kabinetBusyForCount)
pass 
kabinetList = list(kabinetSet)
if 'Blackboard' in kabinetList:
    kabinetList.remove('Blackboard')
kabinetList.sort()
dataDayWeekList = list(dataDayWeekSet)
dataDayWeekList.sort()

for i in dataDayWeekList:
    dataList.append(i[:i.find('\n')])
    dayWeekList.append(i[i.find('\n')+1:])

# Формирование основы таблицы для занятости аудиторий 
kabinetBusyList = list()
rowKabinetList:int = 0
for i in range(0,len(kabinetList)):
    for j in range(0,len(dataList)):
        for k in range(0,len(timeList)):
            kabinetBusyListString = list()
            kabinetBusyListString.append(kabinetList[i])
            kabinetBusyListString.append(dataList[j])
            kabinetBusyListString.append(dayWeekList[j])
            kabinetBusyListString.append(timeList[k])
            kabinetBusyListString.append('Свободна')
            kabinetBusyListString.append('')
            kabinetBusyListString.append('')
            kabinetBusyListString.append('')
            rowKabinetList += 1
            kabinetBusyList.append(kabinetBusyListString)

for i in range(0, len(kabinetList)): # по списку кабинетов
    kbc = kabinetBusyForCount.count(kabinetList[i])
    for j in range(0, kbc): # по расписанию кол-вом записей про кабинет
        kbi = kabinetBusyForCount.index(kabinetList[i])
        sh10 = sheduleList[kbi][10]
        sh4 = sheduleList[kbi][4]
        sh6 = sheduleList[kbi][6]
        for k in range(i*len(dataList)*len(timeList),(i+1)*len(dataList)*len(timeList)): # по занятости кабинета 
            bz0 = kabinetBusyList[k][0] 
            bz1 = kabinetBusyList[k][1] 
            bz3 = kabinetBusyList[k][3] 
            if  bz0 == sh10 and bz1 == sh4 and bz3 == sh6:
                kabinetBusyList[k][4] = 'Занята'
                kabinetBusyList[k][5] = sheduleList[kbi][7]
                kabinetBusyList[k][6] = sheduleList[kbi][8]
                kabinetBusyList[k][7] = sheduleList[kbi][9]
                kabinetBusyForCount[kbi] = kabinetBusyForCount[kbi].lower()
                break

# Работа с openpyxl по сохранению данных в книге XLSX
print()

os.chdir('..\\')
print('Формируем Расписание')
wbXLSX = opx.Workbook()
sheetShedule = wbXLSX.active
sheetShedule.title = 'Расписание'
a = sheetShedule.cell(row=1, column=1, value='Факультет')
b = sheetShedule.cell(row=1, column=2, value='Курс') # Учебный год
c = sheetShedule.cell(row=1, column=3, value='Семестр')
d = sheetShedule.cell(row=1, column=4, value='Неделя')
e = sheetShedule.cell(row=1, column=5, value='Дата')
f = sheetShedule.cell(row=1, column=6, value='День недели')
g = sheetShedule.cell(row=1, column=7, value='Время')
h = sheetShedule.cell(row=1, column=8, value='Группа')
k = sheetShedule.cell(row=1, column=9, value='Дисциплина')
l1 = sheetShedule.cell(row=1, column=10, value='Преподаватель')
m = sheetShedule.cell(row=1, column=11, value='Аудитория')
n = sheetShedule.cell(row=1, column=12, value='Вид занятия')
for i in range(0,len(sheduleList)):
    a = sheetShedule.cell(row=i+2, column=1, value=sheduleList[i][0])
    b = sheetShedule.cell(row=i+2, column=2, value=sheduleList[i][1])
    c = sheetShedule.cell(row=i+2, column=3, value=sheduleList[i][2])
    d = sheetShedule.cell(row=i+2, column=4, value=sheduleList[i][3])
    e = sheetShedule.cell(row=i+2, column=5, value=sheduleList[i][4])
    f = sheetShedule.cell(row=i+2, column=6, value=sheduleList[i][5])
    g = sheetShedule.cell(row=i+2, column=7, value=sheduleList[i][6])
    h = sheetShedule.cell(row=i+2, column=8, value=sheduleList[i][7])
    k = sheetShedule.cell(row=i+2, column=9, value=sheduleList[i][8])
    l2 = sheetShedule.cell(row=i+2, column=10, value=sheduleList[i][9])
    m = sheetShedule.cell(row=i+2, column=11, value=sheduleList[i][10])
    n = sheetShedule.cell(row=i+2, column=12, value=sheduleList[i][11])
print('Формируем Аудитории')
sheetKabinet = wbXLSX.create_sheet(title="Аудитории")

kabinetA = sheetKabinet.cell(row=1, column=1, value='Аудитория')
kabinetB = sheetKabinet.cell(row=1, column=2, value='Дата')
kabinetС = sheetKabinet.cell(row=1, column=3, value='День недели')
kabinetD = sheetKabinet.cell(row=1, column=4, value='Часы')
kabinetE = sheetKabinet.cell(row=1, column=5, value='Занятость')
kabinetF = sheetKabinet.cell(row=1, column=6, value='Группа')
kabinetG = sheetKabinet.cell(row=1, column=7, value='Дисциплина')
kabinetH = sheetKabinet.cell(row=1, column=8, value='Преподаватель')
'''
rowKabinet:int = 2
for i in range(0,len(kabinetList)):
    if kabinetList[i] != 'Blackboard':
        for j in range(0,len(dataList)):
            for k in range(0,len(timeList)):
                kabinetA = sheetKabinet.cell(row=rowKabinet, column=1, value=kabinetList[i])
                kabinetB = sheetKabinet.cell(row=rowKabinet, column=2, value=dataList[j])
                kabinetC = sheetKabinet.cell(row=rowKabinet, column=3, value=dayWeekList[j])
                kabinetD = sheetKabinet.cell(row=rowKabinet, column=4, value=timeList[k])
                rowKabinet += 1
'''
for i in range(len(kabinetBusyList)):
    kabinetA = sheetKabinet.cell(row=i+2, column=1, value=kabinetBusyList[i][0])
    kabinetB = sheetKabinet.cell(row=i+2, column=2, value=kabinetBusyList[i][1])
    kabinetC = sheetKabinet.cell(row=i+2, column=3, value=kabinetBusyList[i][2])
    kabinetD = sheetKabinet.cell(row=i+2, column=4, value=kabinetBusyList[i][3])
    kabinetE = sheetKabinet.cell(row=i+2, column=5, value=kabinetBusyList[i][4])
    kabinetF = sheetKabinet.cell(row=i+2, column=6, value=kabinetBusyList[i][5])
    kabinetG = sheetKabinet.cell(row=i+2, column=7, value=kabinetBusyList[i][6])
    kabinetH = sheetKabinet.cell(row=i+2, column=8, value=kabinetBusyList[i][7])
print('Формируем Преподаватели')
sheetEdukator = wbXLSX.create_sheet(title="Преподаватели")

edukatorList: list =[]
for edukator in edukatorSet:
    edukatorList.append(edukator)
edukatorList.sort()
edukatorA = sheetEdukator.cell(row=1, column=1, value='Преподаватель')
for i in range(0,len(edukatorList)):
    edukatorA = sheetEdukator.cell(row=i+2, column=1, value=edukatorList[i])
ws = wbXLSX['Расписание']
ws.freeze_panes = 'A2'
ws.auto_filter.ref= 'A1:L1'
ws['A1'].font = Font(name = 'Tahoma')
ws['B1'].font = Font(name = 'Tahoma')
ws['C1'].font = Font(name = 'Tahoma')
ws['D1'].font = Font(name = 'Tahoma')
ws['E1'].font = Font(name = 'Tahoma')
ws['F1'].font = Font(name = 'Tahoma')
ws['G1'].font = Font(name = 'Tahoma')
ws['H1'].font = Font(name = 'Tahoma')
ws['I1'].font = Font(name = 'Tahoma')
ws['J1'].font = Font(name = 'Tahoma')
ws['K1'].font = Font(name = 'Tahoma')
ws['L1'].font = Font(name = 'Tahoma')

ws.row_dimensions[1].height = 21
ws.column_dimensions['A'].width = 11
ws.column_dimensions['B'].width = 7 # 12
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 11
ws.column_dimensions['F'].width = 13
ws.column_dimensions['G'].width = 11
ws.column_dimensions['H'].width = 11
ws.column_dimensions['I'].width = 36
ws.column_dimensions['J'].width = 31
ws.column_dimensions['K'].width = 11
ws.column_dimensions['L'].width = 27

ws = wbXLSX['Аудитории']
ws.row_dimensions[1].font = Font(name='Arial', bold=True, size=12)
ws.freeze_panes = 'A2'
ws.auto_filter.ref= 'A1:E1'
ws['A1'].font = Font(name = 'Tahoma')
ws['B1'].font = Font(name = 'Tahoma')
ws['C1'].font = Font(name = 'Tahoma')
ws['D1'].font = Font(name = 'Tahoma')
ws['E1'].font = Font(name = 'Tahoma')
ws.column_dimensions['A'].width = 9
ws.column_dimensions['B'].width = 11
ws.column_dimensions['C'].width = 13
ws.column_dimensions['D'].width = 11
ws.column_dimensions['E'].width = 11
ws.column_dimensions['F'].width = 11
ws.column_dimensions['G'].width = 35
ws.column_dimensions['H'].width = 32

print('Сохраняем')
# ws.row_dimensions[1].font = Font(name='Arial',bold=True, size=12)

wbXLSX.save(fileNameXLSX)

print('ОБРАБОТКА ЗАВЕРШЕНА УСПЕШНО')
# ABCDEFGHIJKLMNOPQRSTUVWXYZ
# maxColumn = sheet.max_column
# maxRow = sheet.max_row
